import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os


class DatabaseIO:
    # Dictionary to store product names and associated filenames
    product_files = {}
    # Static variable to hold the loaded workbook
    wb = None
    sheet = None

    @staticmethod
    def load_database():
        """
        Loads the database into the static variable `wb` and `sheet`.
        This method should be called before any other method that needs access to the workbook.
        """
        database_path = "database.xlsx"

        # Check if the database file exists
        if not os.path.exists(database_path):
            # Create a new workbook and add a sheet
            DatabaseIO.wb = Workbook()
            DatabaseIO.sheet = DatabaseIO.wb.active
            DatabaseIO.sheet.title = "Products"

            # Set column headers
            headers = ["Product", "ID", "Creation Date", "Last Modified"]
            DatabaseIO.sheet.append(headers)

            # Adjust column widths for better display
            column_widths = {
                "A": 50,  # Product column
                "B": 5,  # ID column
                "C": 20,  # Creation Date column
                "D": 20,  # Modified Date column
            }

            for col, width in column_widths.items():
                DatabaseIO.sheet.column_dimensions[col].width = width

            # Save the new workbook to disk
            DatabaseIO.wb.save(database_path)
            print(f"Database created at {database_path}")
        else:
            # If the database exists, load it into the static variables
            DatabaseIO.wb = openpyxl.load_workbook(database_path)
            DatabaseIO.sheet = DatabaseIO.wb["Products"]

    @staticmethod
    def get_available_id():
        """
        Returns the first available ID by checking the existing ones in the database.
        """
        existing_ids = [cell.value for cell in DatabaseIO.sheet["B"][1:]]  # Skip the header row
        new_id = 1

        # Find the first available ID by checking existing ones
        while new_id in existing_ids:
            new_id += 1

        return new_id

    @staticmethod
    def add_product(product_name, filename):
        # Load the database if not already loaded
        if DatabaseIO.wb is None or DatabaseIO.sheet is None:
            DatabaseIO.load_database()

        # Check if the product already exists in the database
        product_found = False
        for row in DatabaseIO.sheet.iter_rows(min_row=2, max_row=DatabaseIO.sheet.max_row, min_col=1, max_col=4):
            if row[0].value == product_name:
                # Check if the creation date already exists
                creation_date_cell = row[0].offset(0, 2)  # Creation Date column (C)
                modified_date_cell = row[0].offset(0, 3)  # Modified Date column (D)

                if creation_date_cell.value is None:  # If no creation date exists, set it
                    creation_date_cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Update the "Modified Date"
                modified_date_cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Update Modified Date
                product_found = True
                print(f"Product '{product_name}' already exists. Modified Date updated.")

                # Add the filename to the dictionary under the product_name
                if product_name in DatabaseIO.product_files:
                    DatabaseIO.product_files[product_name].append(filename)
                else:
                    DatabaseIO.product_files[product_name] = [filename]
                break

        # If product was not found, create a new entry with a new ID
        if not product_found:
            # Get the next available ID using the new method
            new_id = DatabaseIO.get_available_id()

            # Get the current date for creation and modification
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Add the new product with its ID, creation date, and modified date
            DatabaseIO.sheet.append([product_name, new_id, current_date, current_date])
            print(f"Product '{product_name}' with ID {new_id} added to the database.")

            # Add the filename to the dictionary under the product_name
            DatabaseIO.product_files[product_name] = [filename]

        # Save the updated workbook
        DatabaseIO.wb.save("database.xlsx")
